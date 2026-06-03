from pathlib import Path
from typing import Dict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .text_utils import safe_sheet_name


def _is_empty_df(df) -> bool:
    return df is None or not isinstance(df, pd.DataFrame)


def write_dataframes_to_workbook(path: str, sheets: Dict[str, pd.DataFrame]) -> None:
    try:
        import xlsxwriter  # noqa: F401
        return _write_with_xlsxwriter(path, sheets)
    except Exception:
        return _write_with_openpyxl(path, sheets)


def _write_with_xlsxwriter(path: str, sheets: Dict[str, pd.DataFrame]) -> None:
    used_names = set()
    with pd.ExcelWriter(path, engine="xlsxwriter") as writer:
        written = {}
        for name, df in sheets.items():
            if df is None:
                df = pd.DataFrame()
            sheet_name = safe_sheet_name(name, used_names)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            written[sheet_name] = df

        workbook = writer.book
        header_fmt = workbook.add_format({"bold": True, "bg_color": "#D9EAF7", "border": 1})
        money_fmt = workbook.add_format({"num_format": "#,##0.00"})
        int_fmt = workbook.add_format({"num_format": "#,##0"})
        pct_fmt = workbook.add_format({"num_format": "0.00"})

        for sheet_name, worksheet in writer.sheets.items():
            df = written.get(sheet_name, pd.DataFrame())
            row_count = max(len(df), 1)
            col_count = max(len(df.columns), 1)
            worksheet.freeze_panes(1, 0)
            worksheet.autofilter(0, 0, row_count, col_count - 1)
            for col_idx, col_name in enumerate(df.columns):
                width = max(12, min(45, max([len(str(col_name))] + [len(str(v)) for v in df[col_name].head(200).fillna("")]) + 2))
                fmt = None
                col_text = str(col_name)
                if any(k in col_text for k in ["金额", "销售额", "GMV"]):
                    fmt = money_fmt
                elif any(k in col_text for k in ["人数", "次数", "销量", "浏览量", "订单量", "单量", "商品数", "排名"]):
                    fmt = int_fmt
                elif "率" in col_text:
                    fmt = pct_fmt
                worksheet.set_column(col_idx, col_idx, width, fmt)
                worksheet.write(0, col_idx, col_name, header_fmt)


def _write_with_openpyxl(path: str, sheets: Dict[str, pd.DataFrame]) -> None:
    used_names = set()
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            if df is None:
                df = pd.DataFrame()
            df.to_excel(writer, sheet_name=safe_sheet_name(name, used_names), index=False)

    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_column > 1 or ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            values = [ws.cell(row=row_idx, column=col_idx).value for row_idx in range(1, min(ws.max_row, 200) + 1)]
            width = max(12, min(45, max(len(str(v)) if v is not None else 0 for v in values) + 2))
            ws.column_dimensions[letter].width = width
            header = str(ws.cell(row=1, column=col_idx).value or "")
            if any(k in header for k in ["金额", "销售额", "GMV"]):
                number_format = "#,##0.00"
            elif any(k in header for k in ["人数", "次数", "销量", "浏览量", "订单量", "单量", "商品数", "排名"]):
                number_format = "#,##0"
            elif "率" in header:
                number_format = "0.00"
            else:
                number_format = None
            if number_format:
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).number_format = number_format
    wb.save(path)


def build_workbook_sheets(*parts: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
    merged = {}
    for part in parts:
        merged.update(part)
    return merged
